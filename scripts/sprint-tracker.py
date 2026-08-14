#!/usr/bin/env python3
"""Build the client sprint tracker from TASKS.md.

The tracker the client reads is generated from the ledger rather than written by
hand, so it cannot drift from it. The column set is fixed by the tracker he
already keeps: # / Task / Lane / Status / Priority / Sprint / Expected live / done.

    python3 scripts/sprint-tracker.py --sprint "S7 (Fri 8 Aug - Fri 14 Aug)"

Writes SBD_Sprint_Tracker_<today>.xlsx with two sheets. "Current" carries
everything still open plus what shipped inside the recent window, which is what
he actually reads. "Shipped earlier" carries the rest of the done work so the
full list is there without burying the current picture.

Two things this deliberately does not guess. An open task is reported as
"Not started" unless the ledger says otherwise, because the ledger does not
reliably record what is in flight; mark those by hand or fix the ledger. And
priority falls back to Medium only when the entry states none.
"""
import argparse
import datetime
import io
import re
import sys
from pathlib import Path

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
HEADERS = ['#', 'Task', 'Lane', 'Status', 'Priority', 'Sprint', 'Expected live / done']
LANE_ORDER = {'Shipped + verified': 0, 'In progress': 1, 'Open': 2, 'Blocked': 3}
PRIORITY_ORDER = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}

TASK_LINE = re.compile(r'^(  )?- \[([ x])\] (?:~~)?\*\*(T\d+[a-z]?)\*\*(?:~~)? ?(.*)$')
HEADING = re.compile(r'^#{1,3} ')


def parse_tasks(text):
    """Ledger to task records. Sub-tasks are indented and are folded into their parent."""
    tasks, current = [], None
    for line in text.split('\n'):
        m = TASK_LINE.match(line)
        if m:
            if current:
                tasks.append(current)
            current = {'sub': bool(m.group(1)), 'done': m.group(2) == 'x',
                       'id': m.group(3), 'title': m.group(4).replace('~~', ''), 'body': []}
        elif current is not None:
            if HEADING.match(line):
                tasks.append(current)
                current = None
            else:
                current['body'].append(line)
    if current:
        tasks.append(current)
    return [t for t in tasks if not t['sub']]


def clean_title(title):
    """Client-facing wording: drop the estimate tail, internal issue codes and code ticks."""
    title = re.sub(r'\s*·\s*est\b.*$', '', title)
    title = re.sub(r'\s*\(issue\s*`[^`]*`\)', '', title)
    return title.replace('`', '').strip().rstrip('·').strip()


def priority(title):
    """The trailing `· Priority`, whether or not it is bolded.

    Requiring the asterisks silently demoted every unbolded Low and High to
    Medium, which is exactly the column the client sorts on.
    """
    m = re.search(r'·\s*\*{0,2}(Critical|High|Medium|Low)\*{0,2}\s*$', title.strip())
    if m:
        return m.group(1)
    for p in ('Critical', 'High', 'Medium', 'Low'):
        if re.search(r'·\s*\*{0,2}' + p + r'\*{0,2}\b', title):
            return p
    return 'Medium'


def done_date(body):
    m = (re.search(r'\*\*Done (\d{4})-(\d{2})-(\d{2})', body)
         or re.search(r'`done (\d{4})-(\d{2})-(\d{2})', body))
    return tuple(int(g) for g in m.groups()) if m else None


def classify(task):
    """Lane, status and the expected column, from what the entry actually says."""
    body = '\n'.join(task['body'])
    if task['done']:
        d = done_date(body)
        return 'Shipped + verified', 'Done', (f'Live {d[2]} {MONTHS[d[1] - 1]}' if d else 'Live')
    # The ledger marks work that belongs to someone else's queue rather than this one. The
    # marker used to name who; it does not any more, because the client asked to read this
    # ledger directly and who is holding a task is not his business, only that it is held.
    owner = re.search(r'\*Owner: assigned elsewhere[^*]*\*', body)
    if owner:
        return 'In progress', 'In progress', 'Assigned, in progress'
    blocked = re.search(r'\*Blocked on:\*\s*([^\n.]{0,50})', body)
    if blocked:
        return 'Blocked', 'Waiting on ' + blocked.group(1).strip().rstrip('.'), 'Waiting on a decision'
    return 'Open', 'Not started', 'Not scheduled'


def sort_key(row):
    n = int(re.sub(r'\D', '', row['id']))
    return (LANE_ORDER[row['lane']], PRIORITY_ORDER[row['priority']], n, row['id'])


def build_rows(tasks):
    rows = []
    for t in tasks:
        lane, status, expected = classify(t)
        rows.append({'id': t['id'][1:],  # T84a -> 84a, the letter is part of the number
                     'task': clean_title(t['title']),
                     'lane': lane, 'status': status,
                     'priority': priority(t['title']),
                     'expected': expected,
                     'done_date': done_date('\n'.join(t['body']))})
    return rows


def write_workbook(path, current, earlier, sprint):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def fill(ws, rows):
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(1, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='0F2340')
            cell.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append([r['id'], r['task'], r['lane'], r['status'],
                       r['priority'], sprint, r['expected']])
        for i, w in enumerate([6, 78, 20, 26, 10, 26, 24], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=(cell.column == 2), vertical='top')
        ws.freeze_panes = 'A2'

    sheet = wb.active
    sheet.title = 'Current'
    fill(sheet, current)
    fill(wb.create_sheet('Shipped earlier'), earlier)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sprint', required=True, help='e.g. "S7 (Fri 8 Aug - Fri 14 Aug)"')
    ap.add_argument('--tasks', default='TASKS.md')
    ap.add_argument('--recent-from', default='2026-08-01',
                    help='shipped on or after this date stays on the Current sheet')
    ap.add_argument('--out', default=None)
    args = ap.parse_args()

    text = io.open(args.tasks, encoding='utf-8').read()
    rows = build_rows(parse_tasks(text))
    cutoff = tuple(int(x) for x in args.recent_from.split('-'))

    current = [r for r in rows
               if r['lane'] != 'Shipped + verified'
               or (r['done_date'] and r['done_date'] >= cutoff)]
    earlier = [r for r in rows if r not in current]
    current.sort(key=sort_key)
    earlier.sort(key=lambda r: int(re.sub(r'\D', '', r['id'])))

    out = args.out or f'SBD_Sprint_Tracker_{datetime.date.today()}.xlsx'
    write_workbook(out, current, earlier, args.sprint)

    print(f'{out}')
    print(f'  Current         {len(current):>4} rows')
    print(f'  Shipped earlier {len(earlier):>4} rows')
    counts = {}
    for r in current:
        counts[r['lane']] = counts.get(r['lane'], 0) + 1
    for lane in sorted(counts, key=lambda k: LANE_ORDER[k]):
        print(f'    {lane:<20} {counts[lane]}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
